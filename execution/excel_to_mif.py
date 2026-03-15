#!/usr/bin/env python3
"""Excel FRU List → FrameMaker MIF converter.

Replicates VBA macro ButtonGenerateK3_2_Klepnut / ButtonGenerateK3_All_Klepnut.
Always operates in L2_Checked=True mode.

Usage:
    python execution/excel_to_mif.py                        # batch: all pending K3List items
    python execution/excel_to_mif.py --all                  # same
    python execution/excel_to_mif.py <input.xlsx> [template.mif] [output.mif]
"""

import sys
import os
import re
import time
import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
XLSM_FILE = os.path.join(PROJECT_DIR, "Ch3 Generator.xlsm")
TEMPLATE_MIF = os.path.join(PROJECT_DIR, "Level 2 - FRU Master List_templ.mif")
OUTPUT_DIR = os.path.join(PROJECT_DIR, "L2 update")
TBL_SHEET = "TblInsertNewTempl"
GENCH3_SHEET = "GenCh3"
INJECT_DIR = os.path.join(SCRIPT_DIR, 'inject')

# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def col_num(letter):
    return ord(letter.upper()) - ord('A') + 1


def get_contiguous(ws, col, start=2):
    """Rows from `start` until first None (like VBA xlDown)."""
    result = []
    for row in range(start, ws.max_row + 1):
        v = ws.cell(row=row, column=col).value
        if v is None:
            break
        result.append(str(v))
    return result


def get_rows(ws, col, r1, r2):
    """All non-None cell values in rows r1..r2 (inclusive)."""
    result = []
    for row in range(r1, r2 + 1):
        v = ws.cell(row=row, column=col).value
        if v is not None:
            result.append(str(v))
    return result


def strip_apos(s):
    """Excel strips a leading apostrophe from cell values (text-prefix).
    VBA strings like "'  <Para " are stored as "  <Para ".
    """
    return s[1:] if s.startswith("'") else s


def split_200(text):
    """Split text into ≤200-char chunks."""
    chunks = []
    for i in range(0, max(len(text), 1), 200):
        chunks.append(text[i:i+200])
    return chunks


# ---------------------------------------------------------------------------
# Chapter injection helpers
# ---------------------------------------------------------------------------

def load_inject_mif(path):
    """Parse a full or stripped injection .mif file.
    For a full MIF file: extracts all AFrames content + the last TextFlow tagged 'A'.
    Returns (aframes_lines, content_lines).
    """
    with open(path, encoding='latin-1') as f:
        lines = [l.rstrip('\r\n') for l in f]

    aframes = []
    tf_blocks = []       # list of lists — one per TextFlow A block found
    current_tf = None    # lines of the TextFlow block being collected
    current_is_a = False
    mode = None

    for line in lines:
        stripped = line.strip()

        if stripped.startswith('<AFrames'):
            mode = 'aframes'
            continue
        if stripped == '> # end of AFrames':
            mode = None
            continue
        if stripped.startswith('<TextFlow'):
            mode = 'tf'
            current_tf = []
            current_is_a = False
            continue
        if stripped == '> # end of TextFlow':
            if mode == 'tf' and current_is_a:
                tf_blocks.append(current_tf)
            mode = None
            current_tf = None
            continue

        if mode == 'aframes':
            aframes.append(line)
        elif mode == 'tf':
            if stripped.startswith("<TFTag `A'>"):
                current_is_a = True
                continue   # skip TFTag line (header)
            if stripped in ('<TFAutoConnect Yes>', '<Notes', '> # end of Notes'):
                continue   # skip TF header lines
            if stripped.startswith('<TextRectID'):
                continue   # strip page-specific frame refs
            current_tf.append(line)

    content = tf_blocks[-1] if tf_blocks else []
    return aframes, content


def _chapter_name_from_mif(content_lines):
    """Return the first <String `...`> value found in MIF content lines."""
    for line in content_lines:
        s = line.strip()
        if s.startswith("<String `") and s.endswith("'>"):
            return s[9:-2]
    return ''


# ---------------------------------------------------------------------------
# Template blocks
# ---------------------------------------------------------------------------

class Blocks:
    """MIF template blocks from TblInsertNewTempl sheet."""

    def __init__(self, xlsm_path):
        wb     = openpyxl.load_workbook(xlsm_path, data_only=True, read_only=True)
        wb_raw = openpyxl.load_workbook(xlsm_path, read_only=True)
        ws     = wb[TBL_SHEET]
        ws_raw = wb_raw[TBL_SHEET]

        def get_val(row, col, ai_mode=False):
            """Return cell value; formula cells returning None are resolved via raw workbook."""
            v = ws.cell(row=row, column=col).value
            if v is None:
                fv = ws_raw.cell(row=row, column=col).value
                if isinstance(fv, str) and fv.startswith('=IF($B$1="AI",'):
                    if ai_mode:
                        m = re.match(r'^=IF\(\$B\$1="AI","(.*)",".*"\)$', fv)
                        return m.group(1) if m else ''
                    return ''   # formula evaluates to "" when B1 != "AI"
                return None     # truly empty cell
            return str(v)

        def gc(letter, start=2, ai_mode=False):
            col = col_num(letter)
            result = []
            for row in range(start, ws.max_row + 1):
                v = get_val(row, col, ai_mode=ai_mode)
                if v is None:
                    break
                result.append(v)
            return result

        def gr(letter, r1, r2):
            col = col_num(letter)
            result = []
            for row in range(r1, r2 + 1):
                v = get_val(row, col)
                if v is not None:
                    result.append(v)
            return result

        # Before-EOF blocks
        self.I     = gc('I')                 # empty line — non-AI (no font overrides)
        self.I_ai  = gc('I', ai_mode=True)   # empty line — AI section (2pt white)
        self.H     = gc('H', ai_mode=True)   # chapter; build_chapter_block strips FSize/FColor
        self.V     = gc('V', ai_mode=True)   # sub-chapter — always invisible (2pt white)
        self.G     = gc('G')                 # link
        self.J     = gc('J')                 # note
        self.K     = gc('K')                 # caution
        self.U     = gc('U')                 # bullet — non-AI
        self.U_ai  = gc('U', ai_mode=True)   # bullet — AI section (2pt white)
        self.W     = gc('W')                 # text — non-AI
        self.W_ai  = gc('W', ai_mode=True)   # text — AI section (2pt white)
        self.D     = gc('D')                 # ATbl para (standard/consumable/overview)
        self.T     = gc('T')                 # ATbl para (system table)

        # Standard FRU table (col C inserted before > # end of Tbls)
        self.C_full = gc('C')               # full table (header+row+footer)
        self.C_row  = gr('C', 504, 622)     # single row template

        # Consumable table
        self.L_full = gc('L')
        self.L_row  = gr('L', 493, 591)

        # OverView table
        self.N_full = gc('N')
        self.N_row  = gr('N', 504, 622)

        # System table + row
        self.R = gc('R')
        self.S = gc('S')

        wb.close()
        wb_raw.close()


# ---------------------------------------------------------------------------
# MIF document
# ---------------------------------------------------------------------------

class MifDoc:
    """MIF lines stored as list; None = cleared (not written)."""

    def __init__(self, template_path):
        with open(template_path, encoding='latin-1') as f:
            raw = f.read()
        self.lines = [ln.rstrip('\r') for ln in raw.split('\n')]

    def find_first(self, pattern, start=0):
        for i in range(start, len(self.lines)):
            if self.lines[i] is not None and pattern in self.lines[i]:
                return i
        return -1

    def find_last(self, pattern):
        for i in range(len(self.lines) - 1, -1, -1):
            if self.lines[i] is not None and pattern in self.lines[i]:
                return i
        return -1

    def _insert(self, pos, block):
        for j, line in enumerate(block):
            self.lines.insert(pos + j, line)

    def add_before_eof(self, block):
        idx = self.find_last("# End of MIFFile")
        if idx < 0:
            idx = len(self.lines) - 1
        # VBA inserts at lastRow-1 (before '> # end of TextFlow', not before '# End of MIFFile')
        self._insert(idx - 1, block)

    def add_before_end_tbls(self, block):
        """Insert block AT '> # end of Tbls' (shifts it down)."""
        idx = self.find_last("> # end of Tbls")
        self._insert(idx, block)

    def add_row_before_end_tbls(self, block):
        """Insert row template 2 lines before '> # end of Tbls'."""
        idx = self.find_last("> # end of Tbls")
        self._insert(idx - 2, block)

    def add_row_system(self, block):
        """Insert system row 5 lines before '> # end of Tbls'."""
        idx = self.find_last("> # end of Tbls")
        self._insert(idx - 5, block)

    def add_before_end_aframes(self, block):
        idx = self.find_last("# end of AFrames")
        self._insert(idx, block)

    def replace_first(self, what, replacement, start=0):
        idx = self.find_first(what, start)
        if idx >= 0:
            self.lines[idx] = replacement
        return idx

    def change_unique(self):
        counter = 1030000
        for i in range(len(self.lines)):
            line = self.lines[i]
            if line is not None and '<Unique' in line:
                self.lines[i] = re.sub(r'<Unique\s+\d+>', f'<Unique {counter}>', line)
                counter += 1

    def get_output(self):
        return [ln for ln in self.lines if ln is not None and ln != '']


# ---------------------------------------------------------------------------
# Block builders
# ---------------------------------------------------------------------------

def build_chapter_block(H_block, chapter_name):
    """Chapter Para for L2_Checked=True: drop CrossSection variable, set chapter name.
    Strips FSize/FColor so Title 1 chapters are visible in FrameMaker."""
    blk = list(H_block)
    idx = next((i for i, l in enumerate(blk) if l and 'Chapter1Name' in l), -1)
    if idx >= 0:
        for off in (-4, -3, -2, -1):
            t = idx + off
            if 0 <= t < len(blk):
                blk[t] = None
        blk[idx] = f"<String `{chapter_name}'>"
    # Strip FSize/FColor so Title 1 chapter headings are visible in FrameMaker
    blk = [l for l in blk if l is None or ('<FSize' not in l and "<FColor `White'" not in l)]
    return [l for l in blk if l is not None]


def build_subchapter_block(V_block, chapter_name):
    """Sub-chapter Para (AI section): drop CrossSection variable,
    KEEP Pgf font overrides (Title 1.1 AI sub-chapters are intentionally invisible)."""
    blk = list(V_block)
    idx = next((i for i, l in enumerate(blk) if l and 'Chapter1Name' in l), -1)
    if idx >= 0:
        for off in (-4, -3, -2, -1):
            t = idx + off
            if 0 <= t < len(blk):
                blk[t] = None
        blk[idx] = f"<String `{chapter_name}'>"
    return [l for l in blk if l is not None]


def _chunked_block(blk, placeholder, text):
    """Replace placeholder with chunked text (200-char splits) in block."""
    chunks = split_200(text) if text else ['']
    idx = next((i for i, l in enumerate(blk) if l and placeholder in l), -1)
    if idx < 0:
        return [l for l in blk if l is not None]
    blk[idx] = f"<String `{chunks[0]}'>"
    extra = []
    for chunk in chunks[1:]:
        extra += [
            strip_apos("'       > # end of ParaLine"),
            strip_apos("'  <ParaLine "),
            f"<String `{chunk}'>",
        ]
    result = blk[:idx+1] + extra + blk[idx+1:]
    return [l for l in result if l is not None]


def build_note_block(J_block, text):
    blk = [l for l in J_block
           if l is None or ('<FSize' not in l and "<FColor `White'" not in l
                            and '<PgfLeading' not in l)]
    return _chunked_block(blk, 'Chapter1_Note', text)


def build_caution_block(K_block, text):
    blk = [l for l in K_block
           if l is None or ('<FSize' not in l and "<FColor `White'" not in l
                            and '<PgfLeading' not in l)]
    return _chunked_block(blk, 'Chapter1_Caution', text)


def build_bullet_block(U_block, text):
    return _chunked_block(list(U_block), 'Chapter1_Bullet', text)


def build_text_block(W_block, text):
    return _chunked_block(list(W_block), 'Text_Text ', text)


def build_link_block(G_block, link_text, link_hyp, link_free=False):
    blk = list(G_block)
    for i, l in enumerate(blk):
        if l and 'Chapter1_LinkHyp' in l:
            if link_free:
                blk[i] = f"<MText `{link_hyp}'>"
            else:
                blk[i] = f"<MText `openlink ../../database/L3/{link_hyp}.pdf'>"
            break
    for i, l in enumerate(blk):
        if l and 'Chapter1_Link' in l and 'Hyp' not in l:
            blk[i] = f"<String `{link_text}'>"
            break
    return [l for l in blk if l is not None]


# ---------------------------------------------------------------------------
# Remark content builder
# ---------------------------------------------------------------------------

def _hyp_lines(remarkHyp):
    return [
        strip_apos("'        <Font "),
        strip_apos("'         <FTag `Online'>"),
        strip_apos("'         <FUnderlining FSingle>"),
        strip_apos("'         <FLocked No>"),
        strip_apos("'         <FSeparation 4>"),
        strip_apos("'         <FColor `Blue'>"),
        strip_apos("'        > # end of Font"),
        strip_apos("'        <Marker "),
        strip_apos("'         <MType 8>"),
        strip_apos("'         <MTypeName `Hypertext'>"),
        f"         <MText `openlink ../../Database/{remarkHyp}'>",
        strip_apos("'         <MTypeName `Hypertext'>"),
        strip_apos("'         <MCurrPage `1'>"),
        strip_apos("'        > # end of Marker"),
    ]


def _resolve_hyp(raw):
    if '/L2' in raw and raw.endswith('/L2'):
        return f"L2/{raw[:-3]}.fm"
    elif '/L2' in raw:
        return f"L2/{raw}.fm"
    return f"L3/{raw}.pdf"


def build_remark_content(text, hyp_col_f):
    """Build String/Para lines for the remark cell."""
    if not text:
        return ["        <String `'>"]

    has_hyp = '(/' in text
    has_nl = '\n' in text

    hyp_files = [_resolve_hyp(h.strip()) for h in hyp_col_f.split('\n')
                 if h.strip()] if hyp_col_f else []

    if not has_hyp and not has_nl:
        return [f"        <String `{text}'>"]

    out = []
    remark_lines = text.split('\n') if has_nl else [text]
    hyp_num = 0

    for li, rline in enumerate(remark_lines):
        if li > 0:
            out += [
                strip_apos("'       > # end of ParaLine"),
                strip_apos("'      > # end of Para"),
                strip_apos("'      <Para "),
                strip_apos("'       <PgfTag `CellBody'>"),
                strip_apos("'       <ParaLine "),
            ]

        pos = 0
        while pos <= len(rline):
            rem = rline[pos:]
            hp = rem.find('(/')
            if hp < 0:
                if rem:
                    out.append(f"        <String `{rem}'>")
                break
            if hp > 0:
                out.append(f"        <String `{rem[:hp]}'>")
                out += [strip_apos("'       > # end of ParaLine"),
                        strip_apos("'       <ParaLine ")]
                pos += hp
                rem = rline[pos:]
            ep = rem.find('/)')
            if ep < 0:
                out.append(f"        <String `{rem}'>")
                break
            link_text = rem[2:ep]
            rhyp = hyp_files[hyp_num] if hyp_num < len(hyp_files) else 'L3/unknown.pdf'
            hyp_num += 1
            out.extend(_hyp_lines(rhyp))
            out.append(f"        <String `{link_text}'>")
            out += [strip_apos("'        <Font "),
                    strip_apos("'         <FTag `'>"),
                    strip_apos("'         <FLocked No>"),
                    strip_apos("'        > # end of Font")]
            pos += ep + 2

    return out


# ---------------------------------------------------------------------------
# WI helpers
# ---------------------------------------------------------------------------

def _wi_basename(path):
    parts = path.strip('/').split('/')
    return parts[-1] if parts else path


def _wi_extra_paras(wi_extra):
    """Para+Marker+String lines for additional WI references.
    Matches VBA CopyOriginLine sequence where <Unique> ends up after > # end of Marker
    due to VBA inserting at +8 twice (overwriting creates the Unique-after-Marker order).
    """
    return [
        strip_apos("'       <Para "),
        strip_apos("'       <ParaLine "),
        strip_apos("'        <Marker "),
        strip_apos("'         <MType 8>"),
        strip_apos("'         <MTypeName `Hypertext'>"),
        f" <MText `openlink ../../{wi_extra}.fm'>",
        strip_apos("'         <MCurrPage `1'>"),
        strip_apos("'        > # end of Marker"),
        strip_apos("'         <Unique 1032877>"),
        f"<String `{_wi_basename(wi_extra)}'>",
        strip_apos("'       > # end of ParaLine"),
        strip_apos("'      > # end of Para"),
    ]


# ---------------------------------------------------------------------------
# Header substitutions
# ---------------------------------------------------------------------------

def apply_header_subs(doc, doc_number, doc_type, chapter_name,
                      consumable=False, overview=False):
    # doc number
    idx = doc.find_first('xxxxxx')
    if idx >= 0:
        doc.lines[idx] = f" <VariableDef `{doc_number}'>"

    # PDF title
    title = f"{chapter_name} {doc_type} FRU list" if chapter_name else f"{doc_type} FRU list"
    idx = doc.find_first("<Value `FRU list'>")
    if idx >= 0:
        doc.lines[idx] = f" <Value `{title}'>"

    # String ` FRU list' occurrences (for PDF/book title blocks)
    if consumable:
        for _ in range(3):
            idx = doc.find_first("<String ` FRU list'>")
            if idx >= 0:
                doc.lines[idx] = "<String ` List'>"
    elif overview:
        idx = doc.find_first("<String ` FRU list'>")
        if idx >= 0:
            doc.lines[idx] = "<String ` Overview'>"
        for _ in range(2):
            idx = doc.find_first("<String ` FRU list'>")
            if idx >= 0:
                doc.lines[idx] = "<String `'>"

    # CrossSection variable
    cross = f"{chapter_name} {doc_type}" if chapter_name else doc_type
    idx = doc.find_first("<FRU cross section\\>")
    if idx >= 0:
        doc.lines[idx] = f" <VariableDef `{cross}'>"

    # L2_Checked=True: delete revision strings (2 each)
    for _ in range(2):
        idx = doc.find_first("<String `Revision '>")
        if idx >= 0:
            doc.lines[idx] = None
    for _ in range(2):
        idx = doc.find_first("<String `A'>")
        if idx >= 0:
            doc.lines[idx] = None

    # System applicability: clear 9 cells (L2_Checked=True path)
    idx = doc.find_first("[System A], [System B]")
    if idx >= 0:
        for off in (-6, -5, -4, -3, -2, -1, 0, 1, 2):
            t = idx + off
            if 0 <= t < len(doc.lines):
                doc.lines[t] = None

    # Book filenames
    for suf in ('APL', 'LOP', 'LOT', 'LOF', 'TOC'):
        idx = doc.find_first(f'Level 2 - FRU Master List{suf}')
        if idx >= 0:
            doc.lines[idx] = f" <FileName `<c\\>{doc_number}{suf}.fm'>"


# ---------------------------------------------------------------------------
# Row filling (in-document, using find_last to target the newest row)
# ---------------------------------------------------------------------------

def fill_last_row(doc, item_num, description, order_code, wi_raw,
                  remark_text, hyp_col_f, consumable, overview, cons_info):
    """Fill placeholder tokens in the most-recently-inserted row template."""

    # Item number
    idx = doc.find_last("<String `SP_'>")
    if idx >= 0:
        doc.lines[idx] = f"<String ` {item_num}'>"

    # Description
    idx = doc.find_last('SP_001_Description')
    if idx >= 0:
        doc.lines[idx] = f"<String ` {description}'>"

    # --- Order code cell ---
    ohyp = doc.find_last('SP_001_OrderHyp')
    ostr = doc.find_last('SP_001_Order')

    if consumable:
        if ostr >= 0:
            doc.lines[ostr] = f"<String `{order_code}'>" if order_code else ''
    elif overview:
        # OverView: col D = order_code in the WI column; col C goes to remark col
        if ohyp >= 0 and order_code:
            doc.lines[ohyp] = f" <MText `openlink ../../Database/L3/{order_code}_FRU.fm'>"
        if ostr >= 0 and order_code:
            doc.lines[ostr] = f"<String `{order_code}'>"
        # Remark col = col_c (passed as order_code here in overview mode, see main)
    else:
        if order_code and ohyp >= 0:
            doc.lines[ohyp] = f" <MText `openlink ../../Database/L3/{order_code}_FRU.fm'>"
            if ostr >= 0:
                doc.lines[ostr] = f"<String `{order_code}'>"
        elif ohyp >= 0:
            for off in (-6, -3, -2, -1, 0, 1, 2, 3):
                t = ohyp + off
                if 0 <= t < len(doc.lines):
                    doc.lines[t] = None
            if ostr >= 0:
                doc.lines[ostr] = ''

    # --- Consumable info cell ---
    if consumable:
        idx = doc.find_last('Cons_011_Info')
        if idx >= 0:
            doc.lines[idx] = f"<String ` {cons_info}'>"

    # --- WI / Replace Instr cell ---
    rhyp = doc.find_last('SP_001_ReplaceHyp')
    rstr = doc.find_last('SP_001_Replac')

    if wi_raw and not consumable and not overview:
        if wi_raw.startswith('/'):
            all_wi = [w.strip() for w in wi_raw[1:].split('\n')]
        else:
            all_wi = [w.strip() for w in wi_raw.split('\n')]
        all_wi = [w for w in all_wi if w]

        wi_first = all_wi[0]
        wi_rest = all_wi[1:]

        if 'http' in wi_first:
            us = wi_first.find('(/')
            ue = wi_first.find('/)')
            url = wi_first[us+2:ue] if us >= 0 and ue > us else wi_first
            if rhyp >= 0:
                doc.lines[rhyp] = f" <MText `message URL {url}'>"
            wi_display = wi_first[:us].strip() if us > 0 else url
        else:
            if rhyp >= 0:
                doc.lines[rhyp] = f" <MText `openlink ../../{wi_first}.fm'>"
            # VBA: display = Right(D, Len(D) - pos_of_2nd_slash)
            # This preserves newlines and all subsequent paths
            d = wi_raw
            s1 = d.find('/')
            s2 = d.find('/', s1 + 1) if s1 >= 0 else -1
            wi_display = d[s2 + 1:] if s2 >= 0 else _wi_basename(wi_first)

        if rstr >= 0:
            doc.lines[rstr] = f"<String `{wi_display}'>"
            # Insert extra WI Para blocks after '> # end of Para' of the first Para
            # VBA: innerCounter = row(SP_001_Replac) + 2, inserts at innerCounter + 1
            # i.e., after > # end of ParaLine (rstr+1) and > # end of Para (rstr+2)
            if wi_rest:
                insert_pos = rstr + 3
                for we in wi_rest:
                    extra = _wi_extra_paras(we)
                    for j, line in enumerate(extra):
                        doc.lines.insert(insert_pos + j, line)
                    insert_pos += len(extra)
    elif not consumable:
        if rhyp >= 0:
            for off in (-3, -2, -1, 0, 1, 2, 3):
                t = rhyp + off
                if 0 <= t < len(doc.lines):
                    doc.lines[t] = None
        if rstr >= 0:
            doc.lines[rstr] = ''

    # --- Remark cell ---
    if not overview:
        ridx = doc.find_last('SP_001_Remark')
        if ridx >= 0:
            rlines = build_remark_content(remark_text or '', hyp_col_f or '')
            if len(rlines) == 1:
                doc.lines[ridx] = rlines[0]
            else:
                doc.lines[ridx:ridx+1] = rlines
    else:
        # OverView: remark column = order code (col C passed via order_code)
        ridx = doc.find_last('SP_001_Remark')
        if ridx >= 0:
            doc.lines[ridx] = f"<String `{order_code}'>" if order_code else ''


# ---------------------------------------------------------------------------
# GenCh3 loader
# ---------------------------------------------------------------------------

def load_gench3(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb[GENCH3_SHEET]

    def cell(row, col):
        v = ws.cell(row=row, column=col).value
        if v is None:
            return ''
        return str(v).strip()

    meta = {
        'A': cell(1, 1),  # doc_number
        'B': cell(1, 2),  # doc_type
        'C': cell(1, 3),  # chapter_name
    }

    last = 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is not None:
            last = row

    rows = []
    for row in range(2, last + 1):
        rows.append({
            'A': cell(row, 1),
            'B': cell(row, 2),
            'C': cell(row, 3),
            'D': cell(row, 4),
            'E': cell(row, 5),
            'F': cell(row, 6),
            'G': cell(row, 7),
            'L': cell(row, 12),
            'M': cell(row, 13),
            'N': cell(row, 14),
        })

    wb.close()
    return meta, rows


# ---------------------------------------------------------------------------
# K3List / K3DataNew loaders  (batch mode)
# ---------------------------------------------------------------------------

def _cell_str(v):
    """Normalise an openpyxl cell value to a plain string."""
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def load_k3list(xlsm_path):
    """Read K3List sheet and return pending 'new templ' items.

    An item is pending when col D = 'Ok' and col E = '' (not 'Done').
    Returns list of dicts: {doc_number, data_sheet}.
    Uses iter_rows for fast sequential reading of the xlsm.
    """
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, read_only=True)
    ws = wb['K3List']
    items = []
    for rv in ws.iter_rows(min_row=2, values_only=True):
        if not rv or not rv[0]:
            break
        d_col    = _cell_str(rv[3] if len(rv) > 3 else None)   # col D
        done_col = _cell_str(rv[4] if len(rv) > 4 else None)   # col E
        template = _cell_str(rv[5] if len(rv) > 5 else None)   # col F
        ds_ovr   = _cell_str(rv[6] if len(rv) > 6 else None)   # col G
        if d_col == 'Ok' and done_col == '' and template == 'new templ':
            ds = ds_ovr if ds_ovr in ('K3DataNew XPS', 'K3DataNew DTS') else 'K3DataNew'
            items.append({'doc_number': _cell_str(rv[0]), 'data_sheet': ds})
    wb.close()
    return items


def load_from_k3data(xlsm_path, data_sheet, doc_number):
    """Load FRU data for doc_number from a K3DataNew-family sheet.

    Builds an in-memory GenCh3-equivalent array from K3DataNew, assuming
    columns I-M have been deleted from the source sheet.  I-M values are
    synthesised: I = doc_number (header col A); L/M/N computed per section.

    Returns (meta, rows) identical in structure to load_gench3().
    """
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, read_only=True)
    ws = wb[data_sheet]
    sheet_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    doc_str = _cell_str(doc_number)

    def gcol(row_tuple, idx, default=''):
        if row_tuple and len(row_tuple) > idx:
            return _cell_str(row_tuple[idx])
        return default

    def _compute_m(b, c, d, e, f):
        """Python equivalent of the column-M Excel formula in K3DataNew sheets."""
        if b == c:                          # System-list "Applicable to" rows
            return f'Applicable to: [{b}];'
        if b == 'Note':
            return f'Note: [{e.replace(chr(10), " ")}];'
        if b == 'Caution':
            return f'Caution: [{d.replace(chr(10), " ")}];'
        if b == 'Link':
            e2 = e.replace('(/', '').replace('/)', '').replace(chr(10), ' ')
            f2 = f.replace(chr(10), ', ')
            return f'Link: [{e2}], Hyperlink: [{f2}];'
        c2 = c.replace(chr(10), ' ')
        d2 = (d.replace('Database/L2/', '').replace('Database/L3/', '')
                .replace(chr(10), ', '))
        e2 = e.replace('(/', '').replace('/)', '').replace(chr(10), ' ')
        return (f'Description: [{b}], FRU Code, FRU number, Order Code: [{c2}], '
                f'Replacement Work Instruction: [{d2}], Remarks/Other: [{e2}];')

    def data_row(r, col_l='', col_m='', col_n=''):
        """Build a GenCh3-equivalent row dict from a raw K3DataNew data row."""
        return {
            'A': gcol(r, 0), 'B': gcol(r, 1),
            'C': gcol(r, 2), 'D': gcol(r, 3),
            'E': gcol(r, 4), 'F': gcol(r, 5),
            'G': gcol(r, 6),
            'I': ri,         'J': '',          'K': '',
            'L': col_l,      'M': col_m,       'N': col_n,
        }

    # --- Find header row (VBA: Cells.Find(What:=doc_number)) ---
    header_idx = None
    for i, rv in enumerate(sheet_rows):
        if gcol(rv, 0) == doc_str:
            header_idx = i
            break
    if header_idx is None:
        # Fallback: search all columns
        for i, rv in enumerate(sheet_rows):
            if rv and any(_cell_str(v) == doc_str for v in rv):
                header_idx = i
                break
    if header_idx is None:
        raise ValueError(f"{doc_number!r} not found in sheet '{data_sheet}'")

    # --- Find last data row (VBA: innerCounter_end loop, col A) ---
    end_idx = header_idx + 1
    while end_idx < len(sheet_rows):
        v = gcol(sheet_rows[end_idx], 0)
        if v in ('', 'Not Used'):
            break
        end_idx += 1
    data = sheet_rows[header_idx + 1: end_idx]

    # --- Meta from header row (A=doc_number, B=doc_type, C=ch_meta) ---
    hdr = sheet_rows[header_idx]
    meta = {'A': gcol(hdr, 0), 'B': gcol(hdr, 1), 'C': gcol(hdr, 2)}

    # I = doc_number constant (was K3DataNew col I; now derived from header col A)
    ri = gcol(hdr, 0)

    # --- Main section: VBA copies A:H from K3DataNew; L/M/N empty ---
    rows = [data_row(r) for r in data]

    # --- AI section: always included for K3DataNew when data exists ---
    # (col L was 'Bullet' in K3DataNew before I-M columns were deleted)
    if data:
        svc = {
            'A': 'Service Virtual Assistant Content', 'B': 'Text',
            'C': '', 'D': '',
            'E': 'The following text is used by AI to answer FRU list related questions more accurately.',
            'F': '', 'G': '', 'I': '', 'J': '', 'K': '', 'L': '', 'M': '', 'N': '',
        }
        rows.append(dict(svc))                       # service row 1

        for r in data:                               # AI first copy: L='Bullet', M=computed
            m = _compute_m(gcol(r, 1), gcol(r, 2),
                           gcol(r, 3), gcol(r, 4), gcol(r, 5))
            rows.append(data_row(r, col_l='Bullet', col_m=m))

        rows.append(dict(svc))                       # service row 2
        rows.append({                                # MarkDown row
            'A': 'MarkDown', 'B': '', 'C': '', 'D': '', 'E': '',
            'F': '', 'G': '', 'I': '', 'J': '', 'K': '', 'L': '', 'M': '', 'N': 'AI',
        })
        for r in data:                               # AI second copy: N='AI'
            rows.append(data_row(r, col_n='AI'))

    return meta, rows


# ---------------------------------------------------------------------------
# Main conversion
# ---------------------------------------------------------------------------

def generate_mif(meta, rows, blocks, template_path, output_path):
    """Core MIF generation from in-memory meta/rows data."""
    doc_number = meta['A']
    doc_type   = meta['B']
    ch_meta    = meta['C']

    consumable = (doc_type == 'Consumable')
    overview   = (doc_type == 'Overview')

    t_start = time.perf_counter()
    print("Preparing template...")
    doc = MifDoc(template_path)
    apply_header_subs(doc, doc_number, doc_type, ch_meta,
                      consumable=consumable, overview=overview)

    # --- Main row loop ---
    tbl_id = 11
    aframe_id = 1
    chapter = ''
    tbl_one = True
    id_count = 1
    tbl_id_ai = 1   # VBA: TblIdAI – counter for N='AI' and AIContent chapter headings
    sys_done = False
    ai_content = False
    inject_name_cache = {}   # col-C path → resolved chapter name

    i = 0
    while i < len(rows):
        r = rows[i]
        ca = r['A']   # chapter / tag
        cb = r['B']   # content type or description
        cc = r['C']   # order code
        cd = r['D']   # WI ref
        ce = r['E']   # remark / note text
        cf = r['F']   # remark hyp file(s)
        cg = r['G']   # item number
        cl = r['L']   # "Bullet" marker
        cm = r['M']   # bullet text
        cn = r['N']   # "AI" marker

        if ca == '' or ca == 'Not Used':
            break

        # MarkDown block
        if ca == 'MarkDown':
            doc.add_before_eof(blocks.I_ai)
            doc.add_before_eof(build_text_block(blocks.W_ai,
                f'# {ch_meta} {doc_type} FRU list'))
            doc.add_before_eof(blocks.I_ai)
            chapter = ''
            ai_content = False
            tbl_id_ai = 1   # VBA resets TblIdAI = 1 at MarkDown
            i += 1
            continue

        # Data-driven chapter injection
        if ca == 'Inject Chapter':
            file_path = cc if os.path.isabs(cc) else os.path.join(INJECT_DIR, cc)
            if cn != 'AI' and not ai_content:
                aframes_lines, content_lines = load_inject_mif(file_path)
                ch_name = (ce if ce and ce != 'Original name'
                           else _chapter_name_from_mif(content_lines))
                inject_name_cache[cc] = ch_name
                if ce and ce != 'Original name':
                    for _j, _l in enumerate(content_lines):
                        s = _l.strip()
                        if s.startswith("<String `") and s.endswith("'>"):
                            indent = _l[: len(_l) - len(_l.lstrip())]
                            content_lines[_j] = indent + f"<String `{ce}'>"
                            break
                doc.add_before_eof(blocks.I)
                doc.add_before_end_aframes(aframes_lines)
                doc.add_before_eof(content_lines)
            elif cn == 'AI':
                ch_name = (ce if ce and ce != 'Original name'
                           else inject_name_cache.get(cc, _chapter_name_from_mif(
                               load_inject_mif(file_path)[1])))
                doc.add_before_eof(blocks.I_ai)
                doc.add_before_eof(build_text_block(blocks.W_ai,
                    f'## {tbl_id_ai}) {ch_name}'))
                doc.add_before_eof(blocks.I_ai)
                tbl_id_ai += 1
            chapter = ''   # reset so the next "Inject Chapter" row also triggers detection
            i += 1
            continue

        # New chapter
        if ca != chapter:
            chapter = ca

            if ca not in ('No Chapter', 'System list'):
                if ai_content or cn == 'AI':
                    doc.add_before_eof(blocks.I_ai)  # 2pt white (AI section)
                else:
                    doc.add_before_eof(blocks.I)     # 12pt black (non-AI)

                if ai_content:
                    doc.add_before_eof(build_subchapter_block(blocks.V, ca))
                    tbl_id_ai += 1
                elif cn == 'AI':
                    # VBA: addTextProc("## " & CStr(TblIdAI) & ") " & ChapterName)
                    ai_ch_name = (ce if ce and ce != 'Original name' else ca)
                    doc.add_before_eof(build_text_block(blocks.W_ai,
                        f'## {tbl_id_ai}) {ai_ch_name}'))
                    doc.add_before_eof(blocks.I_ai)
                    tbl_id_ai += 1
                    i += 1
                    continue
                else:
                    if ca == 'Service Virtual Assistant Content':
                        doc.add_before_eof(build_subchapter_block(blocks.H, ca))
                    else:
                        ch_display = (ce if ce and ce != 'Original name' else ca)
                        doc.add_before_eof(build_chapter_block(blocks.H, ch_display))

                if ca == 'Service Virtual Assistant Content':
                    ai_content = True

                # VBA: TblOne=True / IDCount=1 only for non-No Chapter / non-System list
                tbl_one = True
                id_count = 1

        # === Special content rows ===

        # Image / Chapter / Note / Text / Bullet / AI
        if cb in ('Image', 'Chapter', 'Note', 'Text') or cl == 'Bullet' or cn == 'AI':
            if cl == 'Bullet':
                if cb not in ('Image', 'Chapter'):
                    doc.add_before_eof(build_bullet_block(
                        blocks.U_ai if ai_content else blocks.U, cm))
                i += 1
                continue

            if cn == 'AI':
                if cb not in ('Image', 'Chapter'):
                    if cb == 'Note':
                        doc.add_before_eof(build_text_block(blocks.W_ai,
                            f'**Note**: {ce.replace(chr(10)," ")}'))
                    elif ca == 'System list':
                        doc.add_before_eof(build_text_block(blocks.W_ai,
                            f'**Applicable to**: {cb}'))
                    elif cb == 'Caution':
                        doc.add_before_eof(build_text_block(blocks.W_ai,
                            f'**Caution**: {ce.replace(chr(10)," ")}'))
                    elif cb == 'Link':
                        doc.add_before_eof(build_text_block(blocks.W_ai,
                            f'**Link**: {ce.replace(chr(10)," ")}'))
                        doc.add_before_eof(build_text_block(blocks.W_ai,
                            f'**HyperLink**: {cf.replace(chr(10),", ")}'))
                    else:
                        doc.add_before_eof(build_text_block(blocks.W_ai,
                            f'- **Description**: {cb}  '))
                        doc.add_before_eof(build_text_block(blocks.W_ai,
                            ' **Order Code**, **FRU Code**, **FRU number**: '
                            f'{cc.replace(chr(10)," ")}  '))
                        wi_d = cd.replace('Database/L3/','').replace('Database/L2/','')
                        doc.add_before_eof(build_text_block(blocks.W_ai,
                            ' **Replacement WI**, **Replacement Work Instruction**: '
                            f'{wi_d}  '))
                        rem_c = ce.replace('(/','').replace('/)',''  )
                        doc.add_before_eof(build_text_block(blocks.W_ai,
                            f' **Remarks/Other**: {rem_c}  '))
                        doc.add_before_eof(blocks.I_ai)
                i += 1
                continue

            if cb == 'Text':
                doc.add_before_eof(build_text_block(
                    blocks.W_ai if ai_content else blocks.W,
                    f'Note: {ce.replace(chr(10)," ")}'))
                i += 1
                continue

            # Note / Image Description as note
            doc.add_before_eof(build_note_block(blocks.J, ce))
            i += 1
            continue

        # Caution
        if cb == 'Caution' and ca == chapter:
            doc.add_before_eof(build_caution_block(blocks.K, cd))
            i += 1
            continue

        # Link / LinkFree
        if cb in ('Link', 'LinkFree') and ca == chapter:
            doc.add_before_eof(build_link_block(blocks.G, ce, cf,
                                                 link_free=(cb == 'LinkFree')))
            i += 1
            continue

        # System list
        if ca == 'System list':
            if not sys_done:
                doc.add_before_eof(blocks.T)
                doc.add_before_end_tbls(blocks.R)
                doc.replace_first('<TblID_001>', f'<TblID {tbl_id}>')
                doc.replace_first('<Atbl_001>', f'<ATbl {tbl_id}>')
                tbl_id += 1
                sys_done = True
            sys_row = list(blocks.S)
            for j, l in enumerate(sys_row):
                if l and 'System_List' in l:
                    sys_row[j] = f"<String `{cb}'>"
                    break
            doc.add_row_system([l for l in sys_row if l is not None])
            i += 1
            continue

        # Image Description (AFrame) — simplified
        if cb == 'Image Description' and ca == chapter:
            doc.add_before_eof(build_note_block(blocks.J, f'Image: {cd}'))
            aframe_id += 1
            i += 1
            continue

        # === Table rows ===
        if ca == chapter:
            if cb == '' and not consumable:
                i += 1
                continue

            if tbl_one:
                # First row: insert table definition
                doc.add_before_eof(blocks.D)           # ATbl Para before EOF
                if consumable:
                    doc.add_before_end_tbls(blocks.L_full)
                elif overview:
                    doc.add_before_end_tbls(blocks.N_full)
                    # Rename column headers for OverView
                    doc.replace_first("<String `Order Code'>", "<String `Item ID'>")
                    doc.replace_first("<String `Replace Instr.'>", "<String `Order Code '>")
                    doc.replace_first("<String `Retrofit / Remarks / Other'>",
                                      "<String  `Replace Instr.'>")
                else:
                    doc.add_before_end_tbls(blocks.C_full)
                    # Rename remarks column for L2_Checked standard table
                    doc.replace_first("<String `Retrofit / Remarks / Other'>",
                                      "<String `Remarks / Other'>")

                doc.replace_first('<TblID_001>', f'<TblID {tbl_id}>')
                doc.replace_first('<Atbl_001>', f'<ATbl {tbl_id}>')
                tbl_id += 1
                tbl_one = False
            else:
                if consumable:
                    doc.add_row_before_end_tbls(list(blocks.L_row))
                elif overview:
                    doc.add_row_before_end_tbls(list(blocks.N_row))
                else:
                    doc.add_row_before_end_tbls(list(blocks.C_row))

            # For OverView, the WI column (col D) holds the order code for the link
            ov_order = cd if overview else cc
            fill_last_row(
                doc,
                item_num=cg,
                description=cb,
                order_code=ov_order,
                wi_raw=cd if not overview else '',
                remark_text=ce,
                hyp_col_f=cf,
                consumable=consumable,
                overview=overview,
                cons_info=cd,
            )

        # Revision check
        if i + 1 < len(rows) and rows[i+1]['A'] == 'Revision':
            break

        i += 1

    t_rows = time.perf_counter()
    print(f"Applying Unique IDs...  (rows: {t_rows - t_start:.1f}s)")
    doc.change_unique()

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    output_lines = doc.get_output()
    print(f"Writing {output_path}  ({len(output_lines)} lines)...")
    with open(output_path, 'w', encoding='latin-1', newline='\n') as f:
        for line in output_lines:
            try:
                f.write(line + '\n')
            except UnicodeEncodeError:
                f.write(line.encode('latin-1', errors='replace').decode('latin-1') + '\n')

    t_done = time.perf_counter()
    print(f"Done.  ({t_done - t_start:.1f}s total)")
    return output_path


def convert(excel_path, template_path=None, output_path=None):
    """Single-file mode: read GenCh3 sheet from xlsx and generate MIF."""
    if template_path is None:
        template_path = TEMPLATE_MIF
    for p, name in [(template_path, 'Template MIF'), (XLSM_FILE, 'Ch3 Generator.xlsm')]:
        if not os.path.exists(p):
            raise FileNotFoundError(f"{name} not found: {p}")

    t0 = time.perf_counter()
    print("Loading template blocks...")
    blocks = Blocks(XLSM_FILE)
    print(f"Template blocks loaded.  ({time.perf_counter() - t0:.1f}s)")

    print(f"Loading {os.path.basename(excel_path)}...")
    meta, rows = load_gench3(excel_path)
    print(f"Data loaded.  ({time.perf_counter() - t0:.1f}s)")

    if output_path is None:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        output_path = os.path.join(OUTPUT_DIR, f"{meta['A']}_d1.mif")

    return generate_mif(meta, rows, blocks, template_path, output_path)


def convert_all(xlsm_path=None, template_path=None, output_dir=None):
    """Batch mode: process all pending K3List items from Ch3 Generator.xlsm.

    Replicates ButtonGenerateK3_All_Klepnut: iterates K3List rows where
    col D='Ok', col E='' (not done), col F='new templ', and generates a
    MIF file for each item using data from the appropriate K3DataNew sheet.
    """
    if xlsm_path is None:
        xlsm_path = XLSM_FILE
    if template_path is None:
        template_path = TEMPLATE_MIF
    if output_dir is None:
        output_dir = OUTPUT_DIR

    for p, name in [(xlsm_path, 'Ch3 Generator.xlsm'), (template_path, 'Template MIF')]:
        if not os.path.exists(p):
            raise FileNotFoundError(f"{name} not found: {p}")

    t_batch = time.perf_counter()
    print("Loading template blocks...")
    blocks = Blocks(xlsm_path)
    t_blocks = time.perf_counter()
    print(f"Template blocks loaded.  ({t_blocks - t_batch:.1f}s)")

    items = load_k3list(xlsm_path)
    if not items:
        print("No pending items in K3List (col D='Ok', col E='', col F='new templ').")
        return []

    print(f"Found {len(items)} item(s) to process: "
          f"{', '.join(it['doc_number'] for it in items)}")

    generated = []
    for idx, item in enumerate(items, 1):
        doc_number = item['doc_number']
        data_sheet = item['data_sheet']
        t_item = time.perf_counter()
        print(f"\n[{idx}/{len(items)}] {doc_number}  (sheet: {data_sheet})")
        try:
            meta, rows = load_from_k3data(xlsm_path, data_sheet, doc_number)
            t_load = time.perf_counter()
            print(f"Data loaded.  ({t_load - t_item:.1f}s)")
            os.makedirs(output_dir, exist_ok=True)
            out = os.path.join(output_dir, f"{doc_number}_d1.mif")
            generate_mif(meta, rows, blocks, template_path, out)
            generated.append(out)
        except Exception:
            import traceback
            traceback.print_exc()

    t_end = time.perf_counter()
    print(f"\nBatch complete. Generated {len(generated)}/{len(items)} file(s).  "
          f"Total: {t_end - t_batch:.1f}s")
    return generated


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]

    if not args or args[0] == '--all':
        # Batch mode: process all pending K3List items
        try:
            convert_all()
        except Exception:
            import traceback
            traceback.print_exc()
            sys.exit(1)
        return

    excel_path    = args[0]
    template_path = args[1] if len(args) > 1 else None
    output_path   = args[2] if len(args) > 2 else None

    if not os.path.exists(excel_path):
        alt = os.path.join(PROJECT_DIR, excel_path)
        if os.path.exists(alt):
            excel_path = alt
        else:
            print(f"File not found: {excel_path}")
            sys.exit(1)

    try:
        convert(excel_path, template_path, output_path)
    except Exception:
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
